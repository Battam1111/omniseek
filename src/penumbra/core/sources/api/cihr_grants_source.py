"""cihr_grants — Canadian CIHR health-research grants, the clinical-AI / health-NLP slice (STRUCTURE).
The third of Canada's three federal research councils (NSERC sciences, SSHRC humanities, CIHR health).

CIHR funds health research, where the NLP/ML-relevant work is clinical NLP, medical machine learning,
health data science, predictive models on EHR/clinical text — a slice the other two councils don't
cover. Open data is bulk per-fiscal-year XLSX (FY2025-26 ~12MB, no DataStore / query API; openpyxl
ingest, NEW for Penumbra). The grant sheet (G&A_S&B) carries ApplicationTitle / Abstract / Keywords /
research categories, so the AI/NLP slice is cleanly extractable.

Acquisition (Penumbra's bulk-file pattern, shared via _bulk_funding):
  * The latest "CIHR Investments YYYY-YY" XLSX url is resolved from CKAN ``package_show`` each refresh
    (read the resource url, never hardcode the UUID).
  * Fetch the static annual XLSX at most monthly (cache_ttl 30d), stream the G&A_S&B sheet (openpyxl
    read_only), keep only is_ai_relevant rows, cache those docs query-independent; a query BM25-filters
    the cached subset.

Structure verified live 2026-06-22 (5 sheets; G&A_S&B = 38 cols incl. ApplicationTitle/Abstract/
Keywords; row1 = Nemer / University of Ottawa). Empty cells arrive as None or the literal "None"
string — both normalized to "". explicit_only: a named-query lookup like nserc_awards / nsf_awards.
"""

from __future__ import annotations

import io
import logging
from typing import Optional

import httpx

from penumbra.core import http
from penumbra.core.normalize import Document
from penumbra.core.sources.api._bulk_funding import UA, BulkFundingBase, is_ai_relevant, year_of

logger = logging.getLogger(__name__)

_PKG = "49edb1d7-5cb4-4fa7-897c-515d1aad5da3"
_PKG_SHOW = f"https://open.canada.ca/data/en/api/3/action/package_show?id={_PKG}"
_DATASET_URL = f"https://open.canada.ca/data/en/dataset/{_PKG}"


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s == "None" else s


class CIHRGrantsAdapter(BulkFundingBase):
    name = "cihr_grants"
    description = (
        "加拿大 CIHR 健康研究经费 — 临床 AI/健康 NLP 切片 (加拿大三大联邦研究局之三: NSERC 理工、SSHRC 人文、"
        "CIHR 健康). NLP/ML 在健康这边以 临床 NLP/医学机器学习/健康数据科学/EHR 预测模型 形式存在, 是另两局"
        "不覆盖的. 开放数据为逐年 bulk XLSX (FY2025-26 ~12MB, 无查询 API; openpyxl 解析). 逐笔奖助: 获奖人 + "
        "机构/系 + 金额(CAD) + program + 主题/类别 + 标题/摘要/关键词. 仅收 AI/ML/NLP 相关切片. 命名 penumbra_fetch."
    )
    explicit_only = "CIHR 加拿大健康经费 临床-AI/NLP 切片 (bulk XLSX, 命名 penumbra_fetch); 月级刷新"
    domains = ["funding"]
    regions = ["ca"]
    modes = ["STRUCTURE"]
    _version = "FY202526"

    def _resolve_xlsx_url(self) -> Optional[str]:
        data = http.get_json(_PKG_SHOW)
        if not isinstance(data, dict):
            return None
        res = (data.get("result") or {}).get("resources") or []
        import re
        inv = [r for r in res if re.search(r"CIHR Investments \d{4}-\d{2}", (r.get("name") or ""))
               and (r.get("format") or "").upper() == "XLSX" and r.get("url")]
        inv.sort(key=lambda r: year_of(r.get("name")), reverse=True)
        return inv[0]["url"] if inv else None

    def _build_subset_docs(self) -> list[Document]:
        url = self._resolve_xlsx_url()
        if not url:
            logger.warning("cihr_grants: could not resolve XLSX url via package_show")
            return []
        try:
            r = httpx.get(url, headers={"User-Agent": UA}, timeout=240, follow_redirects=True)
            r.raise_for_status()
            content = r.content
        except Exception as exc:  # noqa: BLE001 — failure → [] (the contract)
            logger.warning("cihr_grants: XLSX fetch failed: %s", exc)
            return []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            logger.warning("cihr_grants: XLSX parse failed: %s", exc)
            return []
        sheet = next((s for s in wb.sheetnames if "G&A" in s or "S&B" in s),
                     wb.sheetnames[0] if wb.sheetnames else None)
        if not sheet:
            return []
        it = wb[sheet].iter_rows(values_only=True)
        try:
            header = [str(h or "") for h in next(it)]
        except StopIteration:
            return []
        docs: list[Document] = []
        for row in it:
            rowd = {header[i]: _clean(row[i] if i < len(row) else None) for i in range(len(header))}
            if not self._is_relevant(rowd):
                continue
            d = self._row_to_doc(rowd)
            if d is not None:
                docs.append(d)
        logger.info("cihr_grants: built %d health-AI/NLP docs from FY2025-26", len(docs))
        return docs

    @staticmethod
    def _pick(rowd: dict, part: str) -> str:
        """Value of the first header containing ``part`` (the columns are long bilingual names)."""
        part = part.lower()
        for k, v in rowd.items():
            if part in k.lower():
                return v
        return ""

    def _is_relevant(self, rowd: dict) -> bool:
        return is_ai_relevant(self._pick(rowd, "ApplicationTitle"),
                              self._pick(rowd, "ApplicationAbstract"),
                              self._pick(rowd, "ApplicationKeywords"),
                              self._pick(rowd, "AllResearchCategoriesEN"),
                              self._pick(rowd, "PrimaryThemeEN"))

    def _row_to_doc(self, rowd: dict) -> Optional[Document]:
        p = self._pick
        name = (p(rowd, "FirstName") + " " + p(rowd, "FamilyName")).strip()
        title = p(rowd, "ApplicationTitle")
        if not (name or title):
            return None
        inst = p(rowd, "InstitutionPaidNameEN") or p(rowd, "ResearchInstitutionNameEN")
        dept = p(rowd, "ResearchInstitutionDepartment")
        program = p(rowd, "ProgramNameEN")
        amount = p(rowd, "TotalAmountAwarded") or p(rowd, "TotalAmountPaid")
        year = p(rowd, "FiscalYear")
        ref = p(rowd, "FundingReferenceNumber")
        theme = p(rowd, "PrimaryThemeEN")
        cats = p(rowd, "AllResearchCategoriesEN")
        keywords = p(rowd, "ApplicationKeywords")
        abstract = p(rowd, "ApplicationAbstract")
        parts = [f"CIHR 健康研究经费. 获奖人: {name} ({inst}{', ' + dept if dept else ''}). Program: {program}."]
        if theme or cats:
            parts.append(f"主题: {theme}. 研究类别: {cats}.")
        if amount:
            parts.append(f"金额: CAD {amount}.")
        if keywords:
            parts.append(f"关键词: {keywords}.")
        if abstract:
            parts.append(abstract)
        return Document(
            source=self.name,
            source_id=f"cihr:{self._version}:{ref or (name + title)[:48]}",
            url=_DATASET_URL,
            title=(title or f"{program} — {name}")[:140],
            content=" ".join(parts),
            author=name or None,
            tags=["funding", "canada", "cihr", "health"],
            metadata={"recipient": name, "institution": inst, "department": dept,
                      "amount_cad": amount, "program": program, "theme": theme,
                      "categories": cats, "fiscal_year": year, "funding_ref": ref,
                      "keywords": keywords},
        )

    def health_check(self) -> tuple[bool, str]:
        url = self._resolve_xlsx_url()
        if url and url.startswith("http"):
            return True, "OK (latest CIHR Investments XLSX resolved via package_show)"
        return False, "could not resolve the latest XLSX via package_show"


from penumbra.core.fetcher import register_adapter  # noqa: E402

register_adapter(CIHRGrantsAdapter())
